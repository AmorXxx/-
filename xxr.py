import openpyxl
workbook = openpyxl.load_workbook("old.xlsx")
worksheet = workbook.get_sheet_by_name("Sheet1")

newworkbook=openpyxl.Workbook()
newworksheet = workbook.active
newworksheet.title = "Amor"

for i in range(2, 230):
    for t in range(2, 1226):
        content = worksheet.cell(i, t).value
        date = worksheet.cell(i, 1).value
        of = worksheet.cell(1,t).value
        newworksheet.cell(i,3,content)
        newworksheet.cell(i, 2, of)
        newworksheet.cell(i, 1, date)

newworkbook.save(filename='myfile.xlsx')










